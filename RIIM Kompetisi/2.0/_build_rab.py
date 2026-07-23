"""
Build RAB from scratch — purchase model (no rental).
Hardware components classified as bahan baku purwarupa.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy

SRC = "2.0/RAB_RIIM_Kompetisi_2026_Smart_Charging.xlsx"
DST = "2.0/RAB_RIIM_Kompetisi_2026_Optimized.xlsx"

wb = openpyxl.load_workbook(SRC)

# Rename
ws1 = wb["Rincian (th1)"]
ws2 = wb["Rincian (th2)"]

new_title = "Data-Centric AI Pipeline untuk Smart Charging Hub Solar Hybrid: Integritas Data, Hybrid Ensemble Forecasting, dan Energy Management System Adaptif"
new_ketua = "Feddy Setio Pribadi, S.T., M.Kom."

for ws in [ws1, ws2]:
    ws["E3"] = new_title
    ws["E5"] = new_ketua

def set_row(ws, row, vals):
    """Set columns A-M (1-13) for a row"""
    for col_idx, val in enumerate(vals, 1):
        ws.cell(row=row, column=col_idx, value=val)

def clear_row(ws, row):
    for c in range(1, 14):
        ws.cell(row=row, column=c, value=None)

# ══════════════════════════════════════════════
# TAHUN I
# ══════════════════════════════════════════════

# BLP — rows 13-18
blp = [
    (13, 1, "Feddy Setio Pribadi, S.T., M.Kom.",
     "Ketua periset; perancang pipeline AI, data integrity, hybrid ensemble, drift detection, CPTS",
     1, 4, 3600000, "OB", 14400000, 14400000, 0, "Satuan biaya RIIM 2026"),
    (14, 2, "Bagaskoro Saputro, S.Si., M.Cs.",
     "Anggota; IoT-edge, backend, MQTT/Modbus bridge, CSMS, dashboard",
     1, 4, 2400000, "OB", 9600000, 9600000, 0, "Satuan biaya RIIM 2026"),
    (15, 3, "Mario Norman Syah, S.Pd., M.Eng.",
     "Anggota; formulasi EMS, CPTS-EMS, digital twin, uji hardware",
     1, 4, 2400000, "OB", 9600000, 9600000, 0, "Satuan biaya RIIM 2026"),
    (16, 4, "Adhe Lingga Dewi, S.Si., M.Si.",
     "Anggota; data quality, kalibrasi sensor, ANN, dataset curation",
     1, 4, 2400000, "OB", 9600000, 9600000, 0, "Satuan biaya RIIM 2026"),
    (17, 5, "Ir. Turnad Lenggo Ginta, S.T., M.T., Ph.D.",
     "Anggota; kebijakan energi, ESG, techno-economic, hilirisasi",
     1, 4, 2400000, "OB", 9600000, 9600000, 0, "Satuan biaya RIIM 2026"),
]
total_blp = 0
for r, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber in blp:
    set_row(ws1, r, [None, None, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber])
    total_blp += jumlah
set_row(ws1, 18, [None, "Total BLP", None, None, None, None, None, None, None, total_blp, total_blp, 0, None])

# Indikator 1 — rows 20-51
set_row(ws1, 20, [None, "B.1", "Indikator 1",
    "Pengadaan purwarupa dan platform uji smart charging hub",
    "Pembelian komponen utama (PV, BESS, inverter), instrumentasi, dan bahan pendukung purwarupa."])

# B.1.1 — Bahan Baku Purwarupa (komponen utama, bukan sewa)
set_row(ws1, 21, [None, "B.1.1", "Bahan Baku Purwarupa", None, None, None, None, None, None, None, None, None, None])
bbp = [
    (22, 1, "Panel surya monokristalin 370 Wp", "15 pcs × 370 Wp = 5,55 kWp, untuk sumber energi utama", 15, 1, 1350000, "unit", 20250000, 20250000, 0, "Spesifikasi teknis; komponen purwarupa"),
    (23, 2, "BESS LiFePO₄ 10 kWh 48V", "Penyimpanan energi dengan BMS internal", 1, 1, 32000000, "unit", 32000000, 32000000, 0, "Spesifikasi teknis; komponen purwarupa"),
    (24, 3, "Hybrid inverter 5 kW on/off-grid", "Inverter dua arah dengan MPPT, komunikasi Modbus", 1, 1, 14000000, "unit", 14000000, 14000000, 0, "Spesifikasi teknis; komponen purwarupa"),
    (25, 4, "Mounting struktur atap + aksesoris", "Rangka panel surya, bracket, stainless steel", 1, 1, 5000000, "paket", 5000000, 5000000, 0, "Komponen purwarupa"),
    (26, 5, "Kabel PV, konektor MC4, conduit", "Kabel DC, AC, grounding untuk interkoneksi sistem", 1, 1, 4500000, "lot", 4500000, 4500000, 0, "Bahan purwarupa"),
]
for r, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber in bbp:
    set_row(ws1, r, [None, None, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber])
bbp_total = sum(x[8] for x in bbp)

# Clear old rows 27-34 (was belanja modal in original)
for r in range(27, 35):
    clear_row(ws1, r)

set_row(ws1, 34, [None, None, "Subtotal", None, None, None, None, None, None, bbp_total, bbp_total, 0, None])

# B.1.2 — Belanja Modal Instrumentasi (max 10% of Rp250jt = Rp25jt)
set_row(ws1, 35, [None, "B.1.2", "Belanja Modal Instrumentasi", None, None, None, None, None, None, None, None, None, None])
modal = [
    (36, 1, "Industrial IoT gateway", "Akuisisi OCPP/Modbus/MQTT, buffering, watchdog", 2, 1, 1500000, "unit", 3000000, 3000000, 0, "Aset institusi"),
    (37, 2, "Embedded edge controller", "Eksekusi inferensi AI, buffering, store-and-forward", 1, 1, 5000000, "unit", 5000000, 5000000, 0, "Aset institusi"),
    (38, 3, "Energy meter tiga fasa", "Metering daya grid, charger, beban", 2, 1, 1500000, "unit", 3000000, 3000000, 0, "Aset institusi"),
    (39, 4, "Weather station + pyranometer", "Sensor iradiasi, suhu, kelembaban", 1, 1, 5500000, "unit", 5500000, 5500000, 0, "Aset institusi"),
    (40, 5, "Managed switch industri + 4G router", "Segmentasi jaringan OT/IoT + konektivitas", 1, 1, 4000000, "unit", 4000000, 4000000, 0, "Aset institusi"),
    (41, 6, "Edge storage NAS 4TB", "Penyimpanan telemetri, log inferensi, backup", 1, 1, 4000000, "unit", 4000000, 4000000, 0, "Aset institusi"),
]
for r, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber in modal:
    set_row(ws1, r, [None, None, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber])
modal_total = sum(x[8] for x in modal)
if modal_total > 25000000:
    print(f"⚠️ Belanja modal Rp{modal_total:,} > Rp25jt (10%) — perlu dikurangi!")
set_row(ws1, 42, [None, None, "Subtotal", None, None, None, None, None, None, modal_total, modal_total, 0, None])

# B.1.3 — Bahan Habis Pakai & Komponen Pendukung
set_row(ws1, 43, [None, "B.1.3", "Bahan Habis Pakai & Komponen Pendukung", None, None, None, None, None, None, None, None, None, None])
bahan = [
    (44, 1, "Set sensor arus (CT) dan tegangan", "Monitoring arus/tegangan PV, BESS, grid, charger", 6, 1, 300000, "unit", 1800000, 1800000, 0, "Komponen purwarupa"),
    (45, 2, "Sensor temperatur multipoint DS18B20", "Monitoring panel, BESS, enclosure", 8, 1, 100000, "unit", 800000, 800000, 0, "Komponen purwarupa"),
    (46, 3, "RS485–Ethernet + CAN converter", "Integrasi inverter, BMS, meter ke gateway", 4, 1, 350000, "unit", 1400000, 1400000, 0, "Komponen purwarupa"),
    (47, 4, "MCB, MCCB, SPD, fuse, enclosure IP65", "Proteksi dan panel distribusi daya", 1, 1, 5000000, "paket", 5000000, 5000000, 0, "Bahan prototipe"),
    (48, 5, "Terminal, cable gland, wiring, PCB, protoboard", "Perakitan panel dan node sensor", 1, 1, 3500000, "lot", 3500000, 3500000, 0, "Bahan prototipe"),
    (49, 6, "ESP32/Arduino sensor node kit", "Node akuisisi data tambahan", 4, 1, 400000, "set", 1600000, 1600000, 0, "Komponen purwarupa"),
    (50, 7, "Industrial data cable + connector", "Kabel RS485/CAN/Ethernet", 1, 1, 1500000, "lot", 1500000, 1500000, 0, "Bahan prototipe"),
]
# Clear old rows 51-55
for r in range(51, 56):
    clear_row(ws1, r)

for r, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber in bahan:
    set_row(ws1, r, [None, None, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber])
bahan_total = sum(x[8] for x in bahan)
set_row(ws1, 55, [None, None, "Subtotal", None, None, None, None, None, None, bahan_total, bahan_total, 0, None])

indikator1 = bbp_total + modal_total + bahan_total
set_row(ws1, 56, [None, "Total Indikator 1", None, None, None, None, None, None, None, indikator1, indikator1, 0, None])

# Indikator 2 — rows 57-71
set_row(ws1, 57, [None, "B.2", "Indikator 2",
    "Pengembangan pipeline AI, perangkat lunak, pengujian, dan luaran",
    "Pengembangan data integrity layer, hybrid ensemble, CPTS, EMS, dashboard, uji dan dokumentasi."])

# B.2.1 — Jasa Pengembangan Perangkat Lunak
set_row(ws1, 58, [None, "B.2.1", "Jasa Pengembangan Perangkat Lunak", None, None, None, None, None, None, None, None, None, None])
jasa = [
    (59, 1, "Pengembangan data integrity layer", "Data quality scoring, anomaly detection, OCPP/Modbus ingestion", 1, 1, 8000000, "paket", 8000000, 8000000, 0, "HPS jasa teknis"),
    (60, 2, "Pengembangan hybrid ensemble forecasting", "LSTM, TCN, LightGBM, bobot adaptif weighted ensemble", 1, 1, 9000000, "paket", 9000000, 9000000, 0, "HPS jasa teknis"),
    (61, 3, "Pengembangan CPTS dan EMS core", "Cyber-physical trust score, rolling horizon MPC, CPTS-aware optimization", 1, 1, 9000000, "paket", 9000000, 9000000, 0, "HPS jasa teknis"),
    (62, 4, "Pengembangan drift-adaptive retraining", "Monitoring drift, trigger retraining, model versioning", 1, 1, 7000000, "paket", 7000000, 7000000, 0, "HPS jasa teknis"),
    (63, 5, "Pengembangan backend, CSMS, dashboard", "API, database, dashboard real-time, visualisasi CPTS+KPI", 1, 1, 7000000, "paket", 7000000, 7000000, 0, "HPS jasa teknis"),
    (64, 6, "Cloud VPS + object storage (6 bln)", "Server, storage, TLS/VPN, RBAC, backup, audit log", 1, 1, 5000000, "paket", 5000000, 5000000, 0, "HPS layanan cloud"),
]
for r, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber in jasa:
    set_row(ws1, r, [None, None, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber])
jasa_total = sum(x[8] for x in jasa)
set_row(ws1, 65, [None, None, "Subtotal", None, None, None, None, None, None, jasa_total, jasa_total, 0, None])

# B.2.2 — Pengujian, Lapangan, dan Luaran
set_row(ws1, 66, [None, "B.2.2", "Pengujian, Lapangan, dan Luaran", None, None, None, None, None, None, None, None, None, None])
uji = [
    (67, 1, "Uji integrasi + injeksi anomali sintetik", "Validasi integrity layer: packet loss, drift, offset, outlier", 1, 1, 5000000, "paket", 5000000, 5000000, 0, "HPS jasa/pengujian"),
    (68, 2, "Uji kinerja forecasting dan drift detection", "14 skenario (D1-D7 + S0-S6), forecast skill, F1", 1, 1, 4500000, "paket", 4500000, 4500000, 0, "HPS jasa/pengujian"),
    (69, 3, "Site survey, FGD, koordinasi mitra", "Profil beban, survei lokasi, koordinasi mitra Sleman", 1, 1, 4000000, "paket", 4000000, 4000000, 0, "SBM tahun berjalan"),
    (70, 4, "Penyusunan KI, DMP, ICD, model card", "Paten sederhana, data dictionary, SOP, dokumentasi teknis", 1, 1, 4500000, "paket", 4500000, 4500000, 0, "HPS jasa/DJKI"),
]
for r, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber in uji:
    set_row(ws1, r, [None, None, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber])
uji_total = sum(x[8] for x in uji)
set_row(ws1, 71, [None, None, "Subtotal", None, None, None, None, None, None, uji_total, uji_total, 0, None])

indikator2 = jasa_total + uji_total
set_row(ws1, 72, [None, "Total Indikator 2", None, None, None, None, None, None, None, indikator2, indikator2, 0, None])

# Clear old rows 73-84 (old B.2 area) to avoid ghost data
for r in range(73, 85):
    clear_row(ws1, r)

total_blnp = indikator1 + indikator2
total_i = total_blp + total_blnp
# Write totals
ws1["A73"] = "TOTAL BLNP"
ws1.cell(row=73, column=10, value=total_blnp)
ws1.cell(row=73, column=11, value=total_blnp)

ws1["A74"] = "TOTAL I (BIAYA LANGSUNG)"
ws1.cell(row=74, column=10, value=total_i)
ws1.cell(row=74, column=11, value=total_i)

# BTL — keep same structure
total_ii = 12500000
ws1["A75"] = "II."
ws1["B75"] = "BIAYA TIDAK LANGSUNG"
# Rows 76-83 were BTL in original
# Keep BTL row references: monev and evaluasi
clear_row(ws1, 76)
clear_row(ws1, 77)
clear_row(ws1, 78)
clear_row(ws1, 79)
clear_row(ws1, 80)
clear_row(ws1, 81)
clear_row(ws1, 82)
clear_row(ws1, 83)

# Monev
set_row(ws1, 76, [None, "A.", "Monitoring Internal Tahap I",
    "Review metodologi, progres, DMP dan realisasi anggaran", None, None, None, None, None, None, None, None, None])
set_row(ws1, 77, [None, None, 1, "Honor reviewer internal", "Monev tahap I", 2, 2, 1000000, "OJ", 4000000, 4000000, 0, "SBM/internal"])
set_row(ws1, 78, [None, None, 2, "Transportasi + akomodasi reviewer", "Monev tahap I", 1, 1, 1750000, "paket", 1750000, 1750000, 0, "SBM"])
set_row(ws1, 79, [None, None, "Subtotal", None, None, None, None, None, None, 5750000, 5750000, 0, None])

# Evaluasi akhir
set_row(ws1, 80, [None, "B.", "Evaluasi Internal Akhir Periode",
    "Evaluasi TKT, luaran dan kesiapan periode berikutnya", None, None, None, None, None, None, None, None, None])
set_row(ws1, 81, [None, None, 1, "Honor reviewer internal", "Evaluasi akhir periode", 2, 2, 1000000, "OJ", 4000000, 4000000, 0, "SBM/internal"])
set_row(ws1, 82, [None, None, 2, "Transportasi + akomodasi reviewer", "Evaluasi akhir periode", 1, 1, 1750000, "paket", 1750000, 1750000, 0, "SBM"])
set_row(ws1, 83, [None, None, "Subtotal", None, None, None, None, None, None, 5750000, 5750000, 0, None])

ws1["A84"] = "TOTAL II (BTL)"
ws1.cell(row=84, column=10, value=total_ii)
ws1.cell(row=84, column=11, value=total_ii)

ws1["A85"] = "TOTAL BIAYA (I + II)"
grand1 = total_i + total_ii
ws1.cell(row=85, column=10, value=grand1)
ws1.cell(row=85, column=11, value=grand1)

print(f"=== TAHUN I ===")
print(f"BLP:         Rp{total_blp:>10,}")
print(f"BBP:         Rp{bbp_total:>10,}  (PV+BESS+inverter)")
print(f"Modal:       Rp{modal_total:>10,}  (instrumentasi)")
print(f"Bahan:       Rp{bahan_total:>10,}  (pendukung)")
print(f"Indikator 1: Rp{indikator1:>10,}")
print(f"Jasa:        Rp{jasa_total:>10,}  (software)")
print(f"Uji:         Rp{uji_total:>10,}  (testing)")
print(f"Indikator 2: Rp{indikator2:>10,}")
print(f"BLNP:        Rp{total_blnp:>10,}")
print(f"Total I:     Rp{total_i:>10,}")
print(f"BTL:         Rp{total_ii:>10,}")
print(f"GRAND:       Rp{grand1:>10,}")
print(f"BLP%:        {total_blp/grand1*100:.1f}%")

# ══════════════════════════════════════════════
# TAHUN II
# ══════════════════════════════════════════════

# BLP — same
for r, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber in blp:
    set_row(ws2, r, [None, None, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber])
total_blp2 = total_blp  # same
set_row(ws2, 18, [None, "Total BLP", None, None, None, None, None, None, None, total_blp2, total_blp2, 0, None])

# Indikator 1 Th II: Operasi pilot & monitoring
set_row(ws2, 20, [None, "B.1", "Indikator 1",
    "Operasi pilot smart charging hub, monitoring, pemeliharaan",
    "Validasi pipeline dalam operasi nyata, monitoring multi-kondisi, pemeliharaan sistem."])
set_row(ws2, 21, [None, "B.1.1", "Bahan Operasional & Pemeliharaan", None, None, None, None, None, None, None, None, None, None])

# Clear all old data rows
for r in range(22, 85):
    clear_row(ws2, r)

ops = [
    (22, 1, "Bahan operasional pilot (12 bln)", "Kabel, konektor, ATK, label, dokumentasi", 1, 1, 6000000, "paket", 6000000, 6000000, 0, "Bahan habis pakai"),
    (23, 2, "Suku cadang dan spare part", "Fuse, MCB, relay, fan, kabel cadangan", 1, 1, 8000000, "paket", 8000000, 8000000, 0, "Bahan habis pakai"),
    (24, 3, "Sensor dan node IoT pengganti", "Penggantian sensor drift/rusak, node tambahan", 1, 1, 5000000, "paket", 5000000, 5000000, 0, "Bahan habis pakai"),
    (25, 4, "Kalibrasi sensor dan alat ukur", "Kalibrasi ulang sensor, meter, pyranometer", 1, 1, 5000000, "paket", 5000000, 5000000, 0, "HPS jasa kalibrasi"),
    (26, 5, "Sewa alat uji power quality & battery analyzer", "Pengukuran power quality, SOC/SOH, efisiensi", 1, 1, 8000000, "paket", 8000000, 8000000, 0, "HPS laboratorium"),
    (27, 6, "Preventive maintenance sistem (2×)", "Inspeksi, cleaning, torque check, health check", 2, 1, 4000000, "kali", 8000000, 8000000, 0, "HPS jasa teknis"),
]
ops_total = sum(x[8] for x in ops)
for r, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber in ops:
    set_row(ws2, r, [None, None, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber])
subtotal_ops_row = 22 + len(ops)
set_row(ws2, subtotal_ops_row, [None, None, "Subtotal", None, None, None, None, None, None, ops_total, ops_total, 0, None])

# B.1.2 — Cloud & Infrastruktur  
cloud_start = subtotal_ops_row + 1
set_row(ws2, cloud_start, [None, "B.1.2", "Cloud & Infrastruktur", None, None, None, None, None, None, None, None, None, None])
cloud = [
    (cloud_start+1, 1, "Cloud VPS + object storage (12 bln)", "Operasi backend, database, dashboard, backup", 1, 1, 15000000, "paket", 15000000, 15000000, 0, "HPS layanan cloud"),
    (cloud_start+2, 2, "Domain, SSL, CDN, hosting", "Layanan web, dashboard publik", 1, 1, 4000000, "paket", 4000000, 4000000, 0, "HPS layanan"),
    (cloud_start+3, 3, "MQTT broker + data streaming infra", "EMQX/HiveMQ, pipeline data, alerting", 1, 1, 5000000, "paket", 5000000, 5000000, 0, "HPS layanan cloud"),
    (cloud_start+4, 4, "Konektivitas 4G untuk gateway pilot (12 bln)", "Data plan untuk 2 gateway", 2, 12, 150000, "bulan", 3600000, 3600000, 0, "SBM/HP"),
]

cloud_total = sum(x[8] for x in cloud)
for r, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber in cloud:
    set_row(ws2, r, [None, None, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber])
subtotal_cloud_row = (cloud_start + 1) + len(cloud)
set_row(ws2, subtotal_cloud_row, [None, None, "Subtotal", None, None, None, None, None, None, cloud_total, cloud_total, 0, None])

indikator1_2 = ops_total + cloud_total
ind1_row = subtotal_cloud_row + 1
set_row(ws2, ind1_row, [None, "Total Indikator 1", None, None, None, None, None, None, None, indikator1_2, indikator1_2, 0, None])

# Indikator 2 Th II: Diseminasi & luaran akhir
ind2_start = ind1_row + 1
set_row(ws2, ind2_start, [None, "B.2", "Indikator 2",
    "Diseminasi, publikasi, KI, dan luaran akhir riset",
    "Publikasi jurnal, paten, FGD, pelaporan, dan paket replikasi."])
disem_start = ind2_start + 1
set_row(ws2, disem_start, [None, "B.2.1", "Diseminasi & Publikasi", None, None, None, None, None, None, None, None, None, None])
disem = [
    (disem_start+1, 1, "Biaya publikasi jurnal internasional Q2", "OA fee publikasi hasil riset", 1, 1, 12000000, "paket", 12000000, 12000000, 0, "HPS jurnal"),
    (disem_start+2, 2, "Pendaftaran paten sederhana + konsultasi KI", "Biaya pendaftaran dan konsultasi", 1, 1, 6000000, "paket", 6000000, 6000000, 0, "DJKI/HPS"),
    (disem_start+3, 3, "FGD validasi hasil + diseminasi", "Penerimaan pengguna, SOP, transfer pengetahuan", 1, 1, 8000000, "paket", 8000000, 8000000, 0, "SBM/internal"),
    (disem_start+4, 4, "Perjalanan monitoring & pengambilan data (6×)", "Monitoring pilot Sleman, troubleshooting, evaluasi", 6, 1, 3000000, "kali", 18000000, 18000000, 0, "SBM tahun berjalan"),
    (disem_start+5, 5, "Seminar/konferensi nasional", "Presentasi hasil riset", 2, 1, 5000000, "kali", 10000000, 10000000, 0, "SBM"),
]
disem_total = sum(x[8] for x in disem)
for r, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber in disem:
    set_row(ws2, r, [None, None, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber])
subtotal_disem_row = disem_start + 1 + len(disem)
set_row(ws2, subtotal_disem_row, [None, None, "Subtotal", None, None, None, None, None, None, disem_total, disem_total, 0, None])

# B.2.2 — Luaran Akhir
luaran_start = subtotal_disem_row + 1
set_row(ws2, luaran_start, [None, "B.2.2", "Luaran Akhir & Pelaporan", None, None, None, None, None, None, None, None, None, None])
luaran = [
    (luaran_start+1, 1, "Data curation, deposit RIN, model card final", "Dataset v2, metadata, reproducibility package", 1, 1, 7000000, "paket", 7000000, 7000000, 0, "HPS jasa data"),
    (luaran_start+2, 2, "Feasibility dan techno-economic brief", "Analisis biaya, emisi, ESG, strategi hilirisasi", 1, 1, 6000000, "paket", 6000000, 6000000, 0, "HPS jasa konsultan"),
    (luaran_start+3, 3, "Laporan akhir + SOP + paket replikasi", "Dokumentasi akhir, SOP operasi, panduan replikasi", 1, 1, 7000000, "paket", 7000000, 7000000, 0, "HPS jasa"),
    (luaran_start+4, 4, "Penetration test + independent security review", "Audit keamanan gateway, API, cloud, dashboard", 1, 1, 7000000, "paket", 7000000, 7000000, 0, "HPS jasa cybersecurity"),
]
luaran_total = sum(x[8] for x in luaran)
for r, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber in luaran:
    set_row(ws2, r, [None, None, no, nama, justif, vol, freq, harga, sat, jumlah, lpdp, mitra, sumber])
subtotal_luaran_row = luaran_start + 1 + len(luaran)
set_row(ws2, subtotal_luaran_row, [None, None, "Subtotal", None, None, None, None, None, None, luaran_total, luaran_total, 0, None])

indikator2_2 = disem_total + luaran_total
ind2_row = subtotal_luaran_row + 1
set_row(ws2, ind2_row, [None, "Total Indikator 2", None, None, None, None, None, None, None, indikator2_2, indikator2_2, 0, None])

total_blnp2 = indikator1_2 + indikator2_2
total_i2 = total_blp2 + total_blnp2

# Write totals
blnp_row = ind2_row + 1
ws2.cell(row=blnp_row, column=1, value="TOTAL BLNP")
ws2.cell(row=blnp_row, column=10, value=total_blnp2)
ws2.cell(row=blnp_row, column=11, value=total_blnp2)

t1_row = blnp_row + 1
ws2.cell(row=t1_row, column=1, value="TOTAL I (BIAYA LANGSUNG)")
ws2.cell(row=t1_row, column=10, value=total_i2)
ws2.cell(row=t1_row, column=11, value=total_i2)

# BTL
total_ii2 = 12500000
btl_start = t1_row + 1
set_row(ws2, btl_start, [None, "A.", "Monitoring Internal", None, None, None, None, None, None, None, None, None, None])
set_row(ws2, btl_start+1, [None, None, 1, "Honor reviewer + transport + akomodasi", "Monev tahap II", 1, 1, 5750000, "paket", 5750000, 5750000, 0, "SBM"])
set_row(ws2, btl_start+2, [None, None, "Subtotal", None, None, None, None, None, None, 5750000, 5750000, 0, None])
set_row(ws2, btl_start+4, [None, "B.", "Evaluasi Akhir", None, None, None, None, None, None, None, None, None, None])
set_row(ws2, btl_start+5, [None, None, 1, "Honor reviewer + transport + akomodasi", "Evaluasi akhir program", 1, 1, 5750000, "paket", 5750000, 5750000, 0, "SBM"])
set_row(ws2, btl_start+6, [None, None, "Subtotal", None, None, None, None, None, None, 5750000, 5750000, 0, None])

btl_total_row = btl_start + 7
ws2.cell(row=btl_total_row, column=1, value="TOTAL II (BTL)")
ws2.cell(row=btl_total_row, column=10, value=total_ii2)
ws2.cell(row=btl_total_row, column=11, value=total_ii2)

grand_row = btl_total_row + 1
ws2.cell(row=grand_row, column=1, value="TOTAL BIAYA (I + II)")
grand2 = total_i2 + total_ii2
ws2.cell(row=grand_row, column=10, value=grand2)
ws2.cell(row=grand_row, column=11, value=grand2)

print(f"\n=== TAHUN II ===")
print(f"BLP:         Rp{total_blp2:>10,}")
print(f"Operasi:     Rp{ops_total:>10,}")
print(f"Cloud:       Rp{cloud_total:>10,}")
print(f"Indikator 1: Rp{indikator1_2:>10,}")
print(f"Disem:       Rp{disem_total:>10,}")
print(f"Luaran:      Rp{luaran_total:>10,}")
print(f"Indikator 2: Rp{indikator2_2:>10,}")
print(f"BLNP:        Rp{total_blnp2:>10,}")
print(f"Total I:     Rp{total_i2:>10,}")
print(f"BTL:         Rp{total_ii2:>10,}")
print(f"GRAND:       Rp{grand2:>10,}")
print(f"BLP%:        {total_blp2/grand2*100:.1f}%")

print(f"\n=== REKAP 2 TAHUN ===")
print(f"Th I:  Rp{grand1:>10,}")
print(f"Th II: Rp{grand2:>10,}")
print(f"Total: Rp{grand1 + grand2:>10,}")

wb.save(DST)
print(f"\n✅ Saved to {DST}")
